import openpyxl
from openpyxl.styles import PatternFill

# NEW: Import your SDK
from smite2_rh_sdk import Smite2RallyHereSDK

# NEW: Provide predefined (but easily modifiable) lists for sandboxes and KV keys
SANDBOX_NAMES = [
    "Dev",
    "OB2 Cert",
    "OB2 Staging",
    "OB2 PTS",
    "~WARNING OB2 Live"
]

KV_KEYS = [
    "RequiredCLVersion",
    "IsDeserterEnforced",
    "CustomMatchesEnabled",
    "MinimumCustomMatchTeamPlayers",
    "AreDevGameModesEnabled",
    "IsCustomLobbyMapSelectionEnabled",
    "IsCustomLobbySpectatingEnabled",
    "IsJunglePracticeEnabled",
    "JSonConfig.url",
    "AreQuestsEnabled",
    "Announcement.ServerMaintenance",
    "Announcement.ClientUpdate",
    "IsItemStoreUndoEnabled",
    "BackfillPhaseDurationOverride",
    "CustomMatchesSearchLimit",
    "DisplayedCurrencies",
    "game.require_item_vendor",
    "GamesightAPIKey",
    "GamesightAPIUrl",
    "GetGudNPCDataEnabled",
    "hotfix.AllowDailyLoginClaim",
    "hotfix.ArePurchasesEnabled",
    "hotfix.clientmovementauthority",
    "hotfix.debugblockedfires",
    "hotfix.DisablePS5LightShafts",
    "Hotfix.enable",
    "hotfix.gameconfigoverridedebug",
    "hotfix.GameplayCueNotifyTagCheckOnRemove",
    "hotfix.imguiconnect",
    "hotfix.MaxClientGuidRemaps",
    "hotfix.serverfps",
    "HUDABTestingIndex",
    "IsCertEnvironment",
    "IsCombatLogEnabled",
    "IsCustomLobbyVoiceEnabled",
    "IsDebugHUDEnabled",
    "IsEmoteOwnershipRequired",
    "IsGetGudEnabled",
    "IsGlobalInvalidationEnabled",
    "IsGodMasteryTabEnabled",
    "IsHUDEditorEnabled",
    "IsMatchSnapshotEnabled",
    "IsMatchVoiceEnabled",
    "IsPartyVoiceEnabled",
    "IsPlayerProfileEnabled",
    "IsStoreEnabled",
    "IsVoiceEnabled",
    "JSonConfig.config_files",
    "JSonConfig.url",
    "LoginRetries",
    "NotificationsEnabled",
    "ShouldAllowCharacterRecycling",
    "ShowPartyPopupInMatch",
    "SurveyLink"
]

def create_sandbox_kv_report(
    sdk,
    sandbox_names: list = SANDBOX_NAMES,
    kv_keys: list = KV_KEYS,
    output_filename: str = "sandbox_kvs_report.xlsx"
):
    """
    Create an Excel spreadsheet comparing the values of the provided KV keys
    across the provided sandbox names, in the exact order given.
    
    :param sdk: An instance of Smite2RallyHereSDK (pre-initialized)
    :param sandbox_names: A list of sandbox 'short_name' or 'name' strings, in the desired order
    :param kv_keys: A list of KV keys in the desired row order
    :param output_filename: Path/filename for the resulting Excel file
    """
    # 1) Fetch the full list of sandboxes available for the product
    all_sandboxes = sdk.get_all_sandboxes_for_product()  # returns a list of dict
    # Looks something like:
    # [
    #   { "sandbox_id": "...", "name": "Dev", "short_name": "Dev", ... },
    #   { "sandbox_id": "...", "name": "CA1", "short_name": "CA 1", ... },
    #   ...
    # ]

    # 2) Create a mapping from name -> sandbox_id for easy lookup
    #    We'll match on either "name" or "short_name" if needed. 
    sandbox_map = {}
    for sbox in all_sandboxes:
        # NOTE: Adjust logic as needed if your sandbox naming is consistent in `short_name` vs. `name`.
        # We'll just store both keys in the map for convenience:
        if sbox.get("name"):
            sandbox_map[sbox["name"]] = sbox["sandbox_id"]
        if sbox.get("short_name"):
            sandbox_map[sbox["short_name"]] = sbox["sandbox_id"]

    # 3) For each sandbox in the provided order, retrieve the KV data
    #    We'll store in a structure: { sandboxName: { kvKey: kvValue, ... }, ... }
    sandbox_kv_data = {}
    for sb_name in sandbox_names:
        sb_id = sandbox_map.get(sb_name)
        if not sb_id:
            # If we can't find a matching sandbox_id for this name, skip or raise an error
            # For now just store an empty dict
            sandbox_kv_data[sb_name] = {}
            continue

        try:
            kv_list = sdk.get_sandbox_kvs(sb_id)  # returns a list of dicts
            # Example:
            # [
            #   { "sandbox_id": "...", "key": "RequiredCLVersion", "value": "9.12", ... },
            #   { "sandbox_id": "...", "key": "IsDeserterEnforced", "value": "true", ... },
            #   ...
            # ]
            kv_dict = {}
            for kv in kv_list:
                kv_dict[kv["key"]] = kv["value"]
            sandbox_kv_data[sb_name] = kv_dict
        except Exception as e:
            print(f"Warning: Could not fetch KVs for sandbox {sb_name} - {e}")
            sandbox_kv_data[sb_name] = {}

    # 4) Create an Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KVs Comparison"

    # 5) Write the header row with sandbox names
    #    Column A is for "KV Keys", then columns B, C, … are for each sandbox
    ws.cell(row=1, column=1, value="KV Key")
    for col_idx, sb_name in enumerate(sandbox_names, start=2):
        ws.cell(row=1, column=col_idx, value=sb_name)

    # For color-coding
    missing_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Light red
    mismatch_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # Light yellow

    # 6) Fill in the rows for each KV key
    for row_idx, kv_key in enumerate(kv_keys, start=2):
        ws.cell(row=row_idx, column=1, value=kv_key)
        # track values in this row for mismatch highlighting
        row_values = []
        for col_idx, sb_name in enumerate(sandbox_names, start=2):
            val = sandbox_kv_data[sb_name].get(kv_key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            row_values.append(val)
            # If value is missing, color it red
            if val is None:
                cell.fill = missing_fill

        # Optional: highlight cells for a row if there is a mismatch
        # Basic approach: if more than 1 distinct non-empty values in row_values, highlight them
        distinct_vals = [v for v in row_values if v is not None]
        if len(set(distinct_vals)) > 1:
            # highlight row to show there's at least one mismatch
            for col_idx in range(2, 2 + len(sandbox_names)):
                ws.cell(row=row_idx, column=col_idx).fill = mismatch_fill

    # 7) Save the Excel file!
    wb.save(output_filename)

# ------------------------------------------------------------------------------
# Example main function to run the report directly:
# ------------------------------------------------------------------------------
def main():
    """
    Creates the Smite2RallyHereSDK instance (pulling credentials from environment or constructor),
    then generates the KV spreadsheet report in the default file sandbox_kvs_report.xlsx.
    """
    # If your env vars are set for the constructor, you can leave all as None:
    sdk = Smite2RallyHereSDK(
        env_client_id=None,
        env_client_secret=None,
        env_base_url=None,
        dev_client_id=None,
        dev_secret_key=None,
        dev_base_url=None
    )

    # Now call the function to produce the report
    create_sandbox_kv_report(sdk=sdk, output_filename="sandbox_kvs_report.xlsx")

    print("KV Spreadsheet created: sandbox_kvs_report.xlsx")

if __name__ == "__main__":
    main()
